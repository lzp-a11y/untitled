import openpyxl
from openpyxl.styles import PatternFill
workbook = openpyxl.load_workbook(r"D:\00222.xlsx")
sheet = workbook['Sheet1']
green_fill = PatternFill(start_color='00AACF91', end_color='00AACF91', fill_type='solid')


def duiz(max, max2, a, b):
    data = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8, 'I': 9, 'J': 10, 'K': 11,
            'L': 12, 'M': 13, 'N': 14, 'O': 15, 'P': 16, 'Q': 17, 'R': 18, 'S': 19, 'T': 20, 'U': 21,
            'V': 22, 'W': 23, 'X': 24, 'Y': 25, 'z': 26}
    number1 = data[a]
    number2 = data[b]
    for i in range(1, max):
        data1 = sheet.cell(i, number1).value
        yanse1 = sheet.cell(i, number1).fill.fgColor.rgb
        for j in range(1, max2):
            data2 = sheet.cell(j, number2).value
            yanse2 = sheet.cell(j, number2).fill.fgColor.rgb
            if data1 == data2 and data1 != 0 and data2 != 0 and yanse1 != '00AACF91' and yanse2 != '00AACF91'and data1 != None:
                sheet.cell(row=i, column=number1).fill = green_fill
                sheet.cell(row=j, column=number2).fill = green_fill
                break


# duiz(500, 500, 'F', 'T')
duiz(500, 500, 'F', 'P')
workbook.save(r"D:\00222.xlsx")
