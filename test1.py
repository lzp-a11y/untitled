import openpyxl
import time
import xlrd

class operation_excel():
    def __init__(self):
        self.workbook = openpyxl.load_workbook(r"E:\python学习\员工管理系统.xlsx")
        self.sheet = self.workbook['Sheet1']
        self.wb = xlrd.open_workbook(r"E:\python学习\员工管理系统.xlsx")  # 打开文件
        self.sheet2 = self.wb.sheet_by_index(0)  # 获取第一个表单的内容，索引从0开始

    def query(self):
        # 查询员工信息
        time_start = time.time()
        max_row = self.sheet.max_row  # 获取最大行
        for i in range(2, max_row+1):
            name = self.sheet['A{}'.format(i)].value
            age = self.sheet['B{}'.format(i)].value
            sex = self.sheet['C{}'.format(i)].value
            print("姓名：",name, "年龄：", age, "性别：", sex)
        time_end = time.time()
        print("耗时：", time_end-time_start)

    def new_inced(self):
        # 新增员工数据
        name = input("请输入新增员工的姓名：")
        age = int(input("请输入新增员工的年龄："))
        while not 15 <= age <= 70:
            age = int(input("您输入的年龄有误，请重新输入："))
        sex = input("请输入新增员工的性别：")
        while not sex == "男" or sex == "女":
            sex = input("您输入的性别有误，请重新输入：")
        rows = [name, age, sex]
        self.sheet.append(rows)        # 在数据末尾加入数值
        print("员工信息新增成功")
        # data1 = self.sheet.cell(2, 1).value
        # print(data1)
        # col = self.sheet.max_column   # 获取最大列
        # self.sheet['A5'] = 1
        # print(row)
        self.workbook.save(filename=r"E:\python学习\员工管理系统.xlsx")

    def stati(self, name):
        # 计数：统计同名的有几个
        max_row = self.sheet.max_row  # 获取最大行
        count = 0
        row = []
        for i in range(1, max_row+1):
            name_excel = self.sheet['A{}'.format(i)].value
            if name == name_excel:
                count = count + 1
                row.append(i)
        return count, row

    def delete_data(self):
        # 删除员工数据
        name = input("请输入要删除的员工姓名")
        while True:
            count, row = self.stati(name)
            if count == 0:
                name = input("查无此人，请重新输入要删除的员工姓名，或按2退出")
                # count = self.stati(name)
                if name == "2":
                    break
            elif count == 1:
                row2 = row[0]
                age = self.sheet['B{}'.format(row2)].value
                sex = self.sheet['C{}'.format(row2)].value
                if_delete = input("查到该员工信息(姓名：{0} 年龄：{1} 性别：{2})，是否删除？\n 输入1删除，输入2退出:"
                                  .format(name, age, sex))
                if if_delete == "1":
                    self.sheet.delete_rows(idx=row2)
                    print("删除成功")
                    self.workbook.save(filename=r"E:\python学习\员工管理系统.xlsx")
                    break
                else:
                    break
            elif count > 1:
                print("查到姓名为{}的信息,共{}条;信息分别是：".format(name, count))
                for i in row:
                    age = self.sheet['B{}'.format(i)].value
                    sex = self.sheet['C{}'.format(i)].value
                    print("查到该员工信息(姓名：{0} 年龄：{1} 性别：{2})".format(name, age, sex))
                if_delete = int(input("查询到多名同名员工，删除第一位则输入1，删除第二位则输入2，以此类推，退出请输入0，请输入："))
                if if_delete != 0:
                    self.sheet.delete_rows(idx=row[if_delete-1])
                    print("删除成功")
                    self.workbook.save(filename=r"E:\python学习\员工管理系统.xlsx")
                    break
                else:
                    break

    def amend(self):
        # 修改数据
        name = input("请输入要修改的员工姓名")
        while True:
            count, row = self.stati(name)
            if count == 0:
                name = input("查无此人，请重新输入要删除的员工姓名，或按2退出")
                if name == "2":
                    break
            elif count == 1:
                row2 = row[0]
                age = self.sheet['B{}'.format(row2)].value
                sex = self.sheet['C{}'.format(row2)].value
                if_delete = input("查到该员工信息(姓名：{0} 年龄：{1} 性别：{2})，是否修改？\n 输入1修改，输入2退出:"
                                  .format(name, age, sex))
                if if_delete == "1":
                    new_age = int(input("请输入{}修改后的年龄:".format(name)))
                    new_sex = input("请输入{}修改后的性别:".format(name))
                    self.sheet['B{}'.format(row2)] = new_age
                    self.sheet['C{}'.format(row2)] = new_sex
                    print("修改成功")
                    self.workbook.save(filename=r"E:\python学习\员工管理系统.xlsx")
                    break
                else:
                    break
            elif count > 1:
                print("查到姓名为{}的信息,共{}条;信息分别是：".format(name, count))
                for i in row:
                    age = self.sheet['B{}'.format(i)].value
                    sex = self.sheet['C{}'.format(i)].value
                    print("查到该员工信息(姓名：{0} 年龄：{1} 性别：{2})".format(name, age, sex))
                if_delete = int(input("查询到多名同名员工，修改第一位则输入1，修改第二位则输入2，以此类推，退出请输入0，请输入："))
                if if_delete != 0:
                    new_age = int(input("请输入{}修改后的年龄:".format(name)))
                    new_sex = input("请输入{}修改后的性别:".format(name))
                    self.sheet['B{}'.format(row[if_delete-1])] = new_age
                    self.sheet['C{}'.format(row[if_delete-1])] = new_sex
                    print("修改成功")
                    self.workbook.save(filename=r"E:\python学习\员工管理系统.xlsx")
                    break
                else:
                    break

a = operation_excel()
# a.test()
# a.query()
# sheet['A4'] = 2
# for row in sheet.rows:
#     print(row)
# sheet.delete_rows(idx=2)     # 删除第二行
# sheet.insert_rows(idx=2)     # 在第二列插入一列，默认一列
# sheet.insert_cols(idx=2)     # 在第二行插入一行，默认一行
