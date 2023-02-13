import itertools
import pandas as pd
import numpy as np
import xlrd
import openpyxl
workbook = openpyxl.load_workbook(r"E:\Desktop\工作簿1.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row
num_x_list = []
for i in range(2, max_row+1):
    value = sheet["A{}".format(i)].value
    num_x_list.append(value)
value2 = sheet["B2"].value


def get_result(value2, list1):
    result = []
    for i in range(1, len(list1)):
        iter = itertools.combinations(list1, i)
        group_item = list(iter)
        for x in range(1, len(group_item)):
            # print(group_item[x], "之和等于", sum(group_item[x]))
    #         if sum(group_item[x]) == hope:
    #             result.append(group_item[x])
    # print('有', len(result), '种加法组合,得出', hope, '的组合是:')
            if sum(group_item[x]) in range(value2-20, value2+20):
                result.append(group_item[x])
    print('有', len(result), '种加法组合,接近', value2, '的组合是:')
    for i in range(len(result)):
        # print(sum(result[i]), ":", end="")
        for j in range(len(result[i])):
            print(result[i][j], end=' ')
        print("和为", sum(result[i]))


if __name__ == '__main__':
    get_result(value2, num_x_list)


#
# workbook = openpyxl.load_workbook(r"E:\Desktop\工作簿1.xlsx")
# sheet = workbook['Sheet1']
# max_row = sheet.max_row
# num_x_list = []
# for i in range(2, max_row+1):
#     value = sheet["A{}".format(i)].value
#     num_x_list.append(value)
# value2 = sheet["B2"].value
#
#
# def get_result(value2, list1):
#     result = []
#     for i in range(1, len(list1)):
#         iter = itertools.combinations(list1, i)
#         group_item = list(iter)
#         # print(group_item)
#         for x in range(1, len(group_item)):
#             # print(group_item[x], "之和等于", sum(group_item[x]))
#             if sum(group_item[x]) == value2:
#                 result.append(group_item[x])
#     print('有', len(result), '种加法组合,得出', value2, '的组合是:')
#     #         if sum(group_item[x]) in range(value2-20, value2+20):
#     #             result.append(group_item[x])
#     # print('有', len(result), '种加法组合,接近', value2, '的组合是:')
#     for i in range(len(result)):
#         # print(sum(result[i]), ":", end="")
#         for j in range(len(result[i])):
#             print(result[i][j], end=' ')
#         print("和为", sum(result[i]))
#     # print('有', len(result), '这种情况得出结果为', hope, ' 组合是', result)
#
#
# if __name__ == '__main__':
#     get_result(value2, num_x_list)

