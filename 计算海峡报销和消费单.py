"""
计算海峡的报销明细单和消费明细单
"""
import os
import re
import pandas as pd
from openpyxl.styles import PatternFill
import openpyxl


class haixia:

    # 计算海峡报销明细单
    # filename1:资金流水文件名  sheet1： 资金流水文件的表单
    # filename2:林倩名单        sheet2： 林倩名单文件的表单
    # filename3:张怡君名单       sheet3： 张怡君名单文件的表单
    # filename4:要保存的文件名
    def haixia_reimbur_list(self, filename1, filename2, filename3, filename4
                     , sheet1, sheet2, sheet3):
        root_dir = os.path.dirname(__file__)  # 获取当前项目所在目录地址
        excl1 = os.path.join(root_dir, filename1)  # 拼接excel表格地址
        excl2 = os.path.join(root_dir, filename2)
        excl3 = os.path.join(root_dir, filename3)

        df1 = pd.read_excel(excl1, sheet_name=sheet1, skiprows=1)  # 省略指定行数，省略第一行
        df2 = pd.read_excel(excl2, sheet_name=sheet2)
        df3 = pd.read_excel(excl3, sheet_name=sheet3)

        # how 合并方式，left左连接，保留left的全部数据， on： 链接的列属性。
        result = pd.merge(df1, df2.loc[:, ['姓名', '人员类型', '现部门']], how='left', on='姓名')
        result2 = pd.merge(result, df3.loc[:, ['姓名', '身份证', '专业线', '用工性质']], how='left', on='姓名')
        filter1 = (result2["人员类型"] == '紧密型')  # 过滤人员类型列 等于紧密型的数据
        filter2 = (result2["人员类型"] == '客服外包')
        filter3 = (result2["人员类型"] == '营业外包')
        filter4 = (result2["人员类型"] == '政企外包')
        filter5 = (result2["状态"] == '出账')
        result3 = result2.loc[filter1 | filter2 | filter3 | filter4]  # 把筛选后的数据赋值给result3
        result3 = result3.loc[filter5]  # 在result3的基础上进行筛选出账
        result3 = pd.DataFrame(result3)  # 将result3保存成数据框架后才能保存到excel

        date = result3.iloc[0, 10]
        result3['日期2'] = ''  # 18
        result3['时间'] = ''  # 19
        result3['餐类'] = ''  # 20
        result3['补贴'] = ''  # 21
        result3['员工消费金额'] = ''  # 22

        for i in range(0, len(result3.index)):
            date = result3.iloc[i, 10]
            pattern1 = r"(\d{4}-\d{1,2}-\d{1,2})"
            pattern1 = re.compile(pattern1)
            pattern2 = r"(\d{1,2}:\d{1,2}:\d{1,2})"
            pattern2 = re.compile(pattern2)
            str_date1 = pattern1.findall(date)  # 获取分离后的日期2023-01-31
            str_date1 = str_date1[0]
            str_date2 = pattern2.findall(date)  # 获取分离后的时间18:18:13
            str_date2 = str_date2[0]
            result3.iloc[i, 18] = str_date1
            result3.iloc[i, 19] = str_date2

        # 根据时间段区分早中晚餐
        for i in range(0, len(result3.index)):
            date = result3.iloc[i, 19]
            if '07:00:00' < date < '09:00:00':
                result3.iloc[i, 21] = 2
                result3.iloc[i, 20] = '早餐'
            elif '10:00:00' < date < '13:30:00':
                result3.iloc[i, 21] = 6
                result3.iloc[i, 20] = '午餐'
            elif '17:00:00' < date < '19:30:00':
                result3.iloc[i, 21] = 2
                result3.iloc[i, 20] = '晚餐'

        result4 = result3.drop_duplicates(['姓名', '日期2', '餐类'])  # 根据这3个字段去重

        result5 = result4.drop_duplicates(['姓名'])  # 根据名字去重，得到完整的名单
        result5 = result5[['姓名'] + ['身份证'] + ['用工性质'] + ['现部门'] + ['专业线'] + ['员工消费金额']]  # 提取字段到新表格

        # result4是去重后的表格
        # result5是根据名字去重后的完整名单表格
        for i in range(0, len(result5.index)):
            a = []
            name2 = result5.iloc[i, 0]
            for j in range(0, len(result4.index)):
                name1 = result4.iloc[j, 2]
                if name1 == name2:
                    qian = result4.iloc[j, 21]
                    a.append(qian)
            result5.iloc[i, 5] = sum(a)

        result6 = result5[(result5['专业线'] != "产互")]  # result6专业线不等于产互的数据
        result6 = result6.copy()  # 复制result6的数据，在复制的数据上进行操作
        sum1 = result6['员工消费金额'].sum()  # 对员工消费金额这列求和
        result6.loc[1000] = ['合计', '', '', '', '', sum1]
        result7 = result5[(result5['专业线'] == "产互")]  # result7专业线等于产互的数据
        result7 = result7.copy()  # 复制result6的数据，在复制的数据上进行操作
        sum2 = result7['员工消费金额'].sum()  # 对员工消费金额这列求和
        result7.loc[1000] = ['合计', '', '', '', '', sum2]

        # 保存数据到同一个excel的不同不页中
        excl4 = os.path.join(root_dir, filename4)
        new_wb = pd.ExcelWriter(excl4)  # 使用ExcelWriter()可以向同一个excel的不同sheet中写入对应的表格数据
        result6.to_excel(new_wb, sheet_name='非合同制主业', index=False)  # 保存result6的数据到非合同制主业 页
        result7.to_excel(new_wb, sheet_name='非合同制产互', index=False)  # 保存result7的数据到非合同制产互 页
        new_wb.close()  # 直接调用关闭接口就可以了，close方法里面有save保存函数

    # 计算海峡消费明细单
    def haixia_deatcons_list(self, filename1, filename2, filename3, filename4
                     , sheet1, sheet2, sheet3):
        root_dir = os.path.dirname(__file__)  # 获取当前项目所在目录地址
        excl1 = os.path.join(root_dir, filename1)  # 拼接excel表格地址
        excl2 = os.path.join(root_dir, filename2)
        excl3 = os.path.join(root_dir, filename3)

        df1 = pd.read_excel(excl1, sheet_name=sheet1, skiprows=1)  # 省略指定行数，省略第一行
        df2 = pd.read_excel(excl2, sheet_name=sheet2)
        df3 = pd.read_excel(excl3, sheet_name=sheet3)

        # how 合并方式，left左连接，保留left的全部数据， on： 链接的列属性。
        result = pd.merge(df1, df2.loc[:, ['姓名', '人员类型', '现部门']], how='left', on='姓名')
        result2 = pd.merge(result, df3.loc[:, ['姓名', '身份证', '专业线', '用工性质']], how='left', on='姓名')
        filter1 = (result2["人员类型"] == '紧密型')  # 过滤人员类型列 等于紧密型的数据
        filter2 = (result2["人员类型"] == '客服外包')
        filter3 = (result2["人员类型"] == '营业外包')
        filter4 = (result2["人员类型"] == '政企外包')
        filter5 = (result2["状态"] == '出账')
        result3 = result2.loc[(filter1 | filter2 | filter3 | filter4)]  # 把筛选后的数据赋值给data1
        result3 = result3.loc[filter5]  # 在result3的基础上进行筛选出账
        result3 = pd.DataFrame(result3)  # 将result3保存成数据框架后才能保存到excel

        result3['刷卡日期'] = ''  # 18
        result3['时间'] = ''  # 19
        result3['餐点'] = ''  # 20
        result3['早餐补贴金额'] = 0  # 21
        result3['午餐补贴金额'] = 0  # 22
        result3['晚餐补贴金额'] = 0  # 23
        result3['补贴金额'] = ''  # 24
        result3['刷卡时间'] = ''  # 25
        result3['早餐刷卡时间'] = None  # 26
        result3['午餐刷卡时间'] = ''  # 27
        result3['晚餐刷卡时间'] = ''  # 28
        result3['合计'] = ''  # 29

        for i in range(0, len(result3.index)):
            date = result3.iloc[i, 10]
            pattern1 = r"(\d{4}-\d{1,2}-\d{1,2})"
            pattern1 = re.compile(pattern1)
            pattern2 = r"(\d{1,2}:\d{1,2}:\d{1,2})"
            pattern2 = re.compile(pattern2)
            str_date1 = pattern1.findall(date)  # 获取分离后的日期2023-01-31
            str_date1 = str_date1[0]
            str_date2 = pattern2.findall(date)  # 获取分离后的时间18:18:13
            str_date2 = str_date2[0]
            result3.iloc[i, 18] = str_date1
            result3.iloc[i, 25] = str_date2

        # 根据时间段区分早中晚餐
        for i in range(0, len(result3.index)):
            date = result3.iloc[i, 25]
            if '07:00:00' < date < '09:00:00':
                result3.iloc[i, 20] = '早餐'
                result3.iloc[i, 21] = 2
                result3.iloc[i, 26] = date  # 早餐刷卡时间赋值
            elif '10:00:00' < date < '13:30:00':
                result3.iloc[i, 20] = '午餐'
                result3.iloc[i, 22] = 6
                result3.iloc[i, 27] = date  # 午餐刷卡时间赋值
            elif '17:00:00' < date < '19:30:00':
                result3.iloc[i, 20] = '晚餐'
                result3.iloc[i, 23] = 2
                result3.iloc[i, 28] = date  # 晚餐刷卡时间赋值

        result4 = result3.drop_duplicates(['姓名', '刷卡日期', '餐点'])  # 根据这3个字段去重

        # 计算早、中、晚补贴金额合计
        for i in range(0, len(result4.index)):
            result4.iloc[i, 29] = result4.iloc[i, 21] + result4.iloc[i, 22] + result4.iloc[i, 23]
        # 排序
        result4 = result4.copy()
        result4.sort_values("姓名", inplace=True, ascending=False)  # inplace: 原地修改 ascending：升序
        result5 = result4[['姓名'] + ['身份证'] + ['专业线'] + ['刷卡日期'] + ['餐点'] + ['早餐刷卡时间'] + ['早餐补贴金额']
                          + ['午餐刷卡时间'] + ['午餐补贴金额'] + ['晚餐刷卡时间'] + ['晚餐补贴金额'] + ['合计'] + ['用工性质']]  # 提取字段到新表格
        result6 = result5.drop_duplicates(['姓名'])  # result6 是根据姓名去重，得到完整的名单

        # 新建一个result7空表来接收数据
        result7 = pd.DataFrame(columns=["姓名", "身份证", "专业线", "刷卡日期", "餐点", "早餐刷卡时间", "早餐补贴金额",
                                        "午餐刷卡时间", "午餐补贴金额", "晚餐刷卡时间", "晚餐补贴金额", "合计", "用工性质"])

        # result8是提取result5 姓名相同的数据， result9是空的一行，只有姓名、合计和总金额，result7是把8和9的数据进行拼接
        for i in range(0, len(result6)):
            name = result6.iloc[i, 0]  # 姓名
            Professional = result6.iloc[i, 2]  # 专业线
            result8 = result5.loc[result5['姓名'] == name, ["姓名", "身份证", "专业线", "刷卡日期", "餐点", "早餐刷卡时间", "早餐补贴金额",
                                                          "午餐刷卡时间", "午餐补贴金额", "晚餐刷卡时间", "晚餐补贴金额", "合计", "用工性质"]]
            result8.sort_values("刷卡日期", inplace=True, ascending=True)  # inplace: 原地修改 ascending：升序
            sum = result8['合计'].sum()
            result9 = pd.DataFrame(columns=["姓名", "身份证", "专业线", "刷卡日期", "餐点", "早餐刷卡时间", "早餐补贴金额",
                                            "午餐刷卡时间", "午餐补贴金额", "晚餐刷卡时间", "晚餐补贴金额", "合计", "用工性质"],
                                   data=[[name, '合计', Professional, '', '', '', '', '', '', '', '', sum, '']])
            result7 = pd.concat([result7, result8, result9], ignore_index=True)

        result10 = result7[(result7['专业线'] != "产互")]  # result6专业线不等于产互的数据
        result11 = result7[(result7['专业线'] == "产互")]  # result7专业线等于产互的数据

        # 保存数据到同一个excel的不同不页中
        excl4 = os.path.join(root_dir, filename4)
        new_wb = pd.ExcelWriter(excl4)  # 使用ExcelWriter()可以向同一个excel的不同sheet中写入对应的表格数据
        result10.to_excel(new_wb, sheet_name='非合同制主业', index=False)  # 保存result6的数据到非合同制主业 页
        result11.to_excel(new_wb, sheet_name='非合同制产互', index=False)  # 保存result7的数据到非合同制产互 页
        new_wb.close()  # 直接调用关闭接口就可以了，close方法里面有save保存函数

        workbook = openpyxl.load_workbook(excl4)
        sheet = workbook['非合同制主业']
        sheet2 = workbook['非合同制产互']
        green_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        max = sheet.max_row
        max2 = sheet2.max_row
        for i in range(1, max + 1):
            value = sheet.cell(i, 2).value
            if value == "合计":
                for j in range(1, 14):
                    sheet.cell(row=i, column=j).fill = green_fill

        for i in range(1, max2 + 1):
            value = sheet2.cell(i, 2).value
            if value == "合计":
                for j in range(1, 14):
                    sheet2.cell(row=i, column=j).fill = green_fill
        workbook.save(excl4)